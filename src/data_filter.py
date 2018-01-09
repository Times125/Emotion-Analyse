#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""
@Author:lch02
@Time:  2018/1/3 13:08
@Description: 利用训练好的分类器重新筛选原始语料，然后利用原始语料重新训练
"""

import os
import config
import pickle
from multiprocessing import Pool
from openpyxl import load_workbook
from deal_files import text_parse
from deal_features import best_bigram_words_features

__author__ = 'lch02'

def review_filter():
    export_data_f()
    with open(os.path.join(config.test_path, 'my_nb_classifier_mod.pkl'), 'rb') as f:
        classifier = pickle.load(f)
    with open(os.path.join(config.test_path, 'neg_review_txt.pkl'), 'rb') as f:
        neg_review_txt = pickle.load(f)
    with open(os.path.join(config.test_path, 'pos_review_txt.pkl'), 'rb') as f:
        pos_review_txt = pickle.load(f)
    reviews = [neg_review_txt, pos_review_txt]
    pools = Pool()
    for i in range(2):
        pools.apply_async(get_better_reviews, args=(i, reviews[i], classifier,))
    pools.close()
    pools.join()

    print 'review filter done!'


def get_better_reviews(cat, reviews, classifier):
    neg_reviews = []
    neg_reviews_lst = []
    pos_reviews = []
    pos_reviews_lst = []
    if cat == 0:
        for words in reviews:
            wordlst = text_parse(words)
            res = classifier.classify(best_bigram_words_features(wordlst))
            if res == 'neg':
                neg_reviews.append(words)
                neg_reviews_lst.append(wordlst)
        with open(os.path.join(config.test_path, 'neg_reviews.pkl'), 'wb') as f:
            pickle.dump(neg_reviews, f)
            print 'better neg_reviews done! %d' % len(neg_reviews)
        with open(os.path.join(config.test_path, 'neg_review.pkl'), 'wb') as f:
            pickle.dump(neg_reviews_lst, f)
            print 'better neg_reviews_lst done!%d' % len(neg_reviews_lst)
    if cat == 1:
        for words in reviews:
            wordlst = text_parse(words)
            res = classifier.classify(best_bigram_words_features(wordlst))
            if res == 'pos':
                pos_reviews.append(words)
                pos_reviews_lst.append(wordlst)
        with open(os.path.join(config.test_path, 'pos_reviews.pkl'), 'wb') as f:
            pickle.dump(pos_reviews, f)
            print 'better pos_reviews done!%d' % len(pos_reviews)
        with open(os.path.join(config.test_path, 'pos_review.pkl'), 'wb') as f:
            pickle.dump(pos_reviews_lst, f)
            print 'better pos_reviews_lst done!%d' % len(pos_reviews_lst)

"""
将Excle中的数据导出
"""
def export_data_f():
    pool = Pool()
    files = ['Sentiment0.xlsx', 'Sentiment4.xlsx']
    for i in range(2):
        pool.apply_async(deal_doc_f, args=(i, files[i]))
    pool.close()
    pool.join()
    print 'import done!'

def deal_doc_f(cat, fn):
    file_name = os.path.join(config.test_path, fn)
    wb = load_workbook(file_name, read_only=True)
    ws = wb.active
    neg = []
    pos = []
    if cat == 0:
        print 'import neg .... %d' % os.getpid()
        for row in ws.iter_rows('A:B'):
            label = row[0].value
            content_txt = row[1].value
            if content_txt is not None:
                content = text_parse(content_txt)
                if len(content) == 0:
                    continue
                elif label == 0 and len(content) != 0:
                    neg.append(content_txt)
        neg_file = os.path.join(config.test_path, 'neg_review_txt.pkl')  # 消极语料
        with open(neg_file, 'wb') as f:
            pickle.dump(neg, f)
    else:
        print 'import pos .... %d' % os.getpid()
        for row in ws.iter_rows('A:B'):
            label = row[0].value
            content_txt = row[1].value
            if content_txt is not None:
                content = text_parse(content_txt)
                if len(content) == 0:
                    continue
                elif label == 4 and len(content) != 0:
                    pos.append(content_txt)
        pos_file = os.path.join(config.test_path, 'pos_review_txt.pkl')  # 积极语料
        with open(pos_file, 'wb') as f:
            pickle.dump(pos, f)
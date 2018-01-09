#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""
@Author:lch02
@Time:  2017/12/25 14:32
@Description: 
"""
import json
import socket

from deal_files import export_data
from deal_features import *
from my_classifier import *
from openpyxl import load_workbook
from deal_files import text_parse
from data_filter import review_filter

__author__ = 'lch02'

def test_classifier():
    file_name = os.path.join(test_path, 'testdata.xlsx')
    wb = load_workbook(file_name, read_only=True)
    ws = wb.active
    neg = []
    pos = []
    for row in ws.iter_rows('A:B'):
        label = row[0].value
        content = row[1].value
        if content is not None:
            content = text_parse(content)
            if len(content) == 0:
                continue
            elif label == 0 and len(content) != 0:
                neg.append((content, 0))
            elif label == 4 and len(content) != 0:
                pos.append((content, 4))
    pos.extend(neg)
    test = pos
    sum = len(test)
    print len(neg), '----', sum
    corrected = 0.0
    classifier = get_model()
    print 'model loaded!'
    for words, label in test:
        f_dict = best_bigram_words_features(words)
        cl = classifier.classify(f_dict)
        if label == 0 and cl == 'neg':
            corrected += 1
        elif label == 4 and cl == 'pos':
            corrected += 1
    print 'accuracy is %.7f' % (corrected / sum)

def main():
    classifier = get_model()
    print 'load success!'
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(('127.0.0.1', 7004))
    s.listen(10)
    while True:
        try:
            sock, addr = s.accept()
            data = sock.recv(102400)
            data = data.decode('utf-8').encode('utf-8')
            if data is 'auto':
                if len(config.best_words) == 0:
                    config.best_words = pickle.load(open(os.path.join(config.test_path, 'best_feats.pkl'), 'rb'))
                create_classifier(best_bigram_words_features)
                sock.send("complete auto!")
                sock.close()
                continue
            deal = best_words_features(text_parse(data))
            res = classifier.classify(deal)
            if res == 'pos':
                p = 1.0
            else:
                p = 0.0
            res_json = json.dumps(p).encode('utf-8')
            sock.send(res_json)
            sock.close()
        except KeyboardInterrupt:
            print 'Exit'
            exit()
        except:
            sock.send('Error')
            sock.close()
            print 'Error'


if __name__ == '__main__':
    main()
    # export_data()
    # review_filter()
    """
    word_bigram_score_dict = word_bigram_scores()
    config.best_words = get_best_words(word_bigram_score_dict, 5000)
    best_feats_pkl = os.path.join(config.test_path, 'best_feats.pkl')
    with open(best_feats_pkl, 'wb') as f:
        pickle.dump(config.best_words, f)
    create_classifier(best_bigram_words_features)
    """
    # test_classifier()  # NB 0.8189415 # SVM 0.5961003
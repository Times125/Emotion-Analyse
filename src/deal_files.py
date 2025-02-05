#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""
@Author:lch02
@Time:  2017/12/25 14:46
@Description: 处理xlxsw中的文本，将其中的数据（160w条）导出
"""
import os
import re
import pickle

from nltk import regexp_tokenize
from nltk.corpus import stopwords
from config import test_path
from openpyxl import load_workbook
from multiprocessing import Pool

__author__ = 'lch02'

"""
将Excle中的数据导出
"""
def export_data():
    pool = Pool()
    files = ['Sentiment0.xlsx', 'Sentiment4.xlsx']
    for i in range(2):
        pool.apply_async(deal_doc, args=(i, files[i]))
    pool.close()
    pool.join()
    print 'import'

def deal_doc(cat, fn):
    file_name = os.path.join(test_path, fn)
    wb = load_workbook(file_name, read_only=True)
    ws = wb.active
    neg = []
    pos = []
    if cat == 0:
        for row in ws.iter_rows('A:B'):
            label = row[0].value
            content = row[1].value
            if content is not None:
                content = text_parse(content)
                if len(content) == 0:
                    continue
                elif label == 0 and len(content) != 0:
                    neg.append(content)
        neg_file = os.path.join(test_path, 'neg_review.pkl')  # 消极语料
        with open(neg_file, 'wb') as f:
            pickle.dump(neg, f)
    else:
        for row in ws.iter_rows('A:B'):
            label = row[0].value
            content = row[1].value
            if content is not None:
                content = text_parse(content)
                if len(content) == 0:
                    continue
                elif label == 4 and len(content) != 0:
                    pos.append(content)
        pos_file = os.path.join(test_path, 'pos_review.pkl')  # 积极语料
        with open(pos_file, 'wb') as f:
            pickle.dump(pos, f)

"""
文本处理：取词、去停用词等
"""
def text_parse(input_text, language='en'):
    sentence = input_text.strip().lower()
    sentence = re.sub(r'@\s*[\w]+ | ?#[\w]+ | ?&[\w]+; | ?[^\x00-\xFF]+', '', sentence)
    special_tag = set(
        ['.', ',', '#', '!', '(', ')', '*', '`', ':', '"', '‘', '’', '“', '”', '@', '：', '^', '/', ']', '[', ';', '=', '_'])
    pattern = r""" (?x)(?:[a-z]\.)+ 
                  | \d+(?:\.\d+)?%?\w+
                  | \w+(?:[-']\w+)*"""

    word_list = regexp_tokenize(sentence, pattern)
    if language == 'en':
        filter_word = [w for w in word_list if
                       w not in stopwords.words('english') and w not in special_tag]  # 去停用词和特殊标点符号
    return filter_word
